# -*- coding: utf-8 -*-
"""
Excel Dashboard Builder for Paper Trading — V9.2
=================================================
يبني ملف Excel يحتوي على:
  Sheet 1: Active Trades (الصفقات المفتوحة)
  Sheet 2: Closed Trades (الصفقات المغلقة)
  Sheet 3: Performance Stats (الإحصاءات)
  Sheet 4: Per-Stock Performance (أداء كل سهم)
  Sheet 5: Per-Sector Performance (أداء كل قطاع)

يُستخدم من run_all.py:
  from paper_trading_excel import build_dashboard
  build_dashboard()
"""
import json
from pathlib import Path
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

BASE = Path("tadawul_data")
PAPER_DIR = Path("paper_trades")
F_TRADES = BASE / "paper_trades.json"

# Styling
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", start_color="D9E1F2")
WIN_FILL = PatternFill("solid", start_color="C6EFCE")
LOSS_FILL = PatternFill("solid", start_color="FFC7CE")
NEUTRAL_FILL = PatternFill("solid", start_color="FFEB9C")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def _style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER


def _style_subheader(cell):
    cell.fill = SUBHEADER_FILL
    cell.font = Font(bold=True, size=11)
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER


def _autosize_columns(ws, max_width=35):
    for col in ws.columns:
        try:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            adjusted = min(max_length + 2, max_width)
            ws.column_dimensions[col[0].column_letter].width = adjusted
        except Exception:
            pass


def build_active_trades_sheet(ws, active_trades):
    """Sheet 1: الصفقات النشطة"""
    headers = [
        "Trade ID", "Date", "Ticker", "Sector", "Signal Type",
        "Entry", "Stop", "T1", "T2", "Score", "Conf", "Days Open",
        "Current", "Unrealized P&L%", "MFE%", "MAE%", "Status"
    ]
    
    # Title row
    ws["A1"] = "📊 الصفقات النشطة (Active Trades)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"].alignment = CENTER_ALIGN
    
    # Headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        _style_header(cell)
    
    # Data
    for row_idx, t in enumerate(active_trades, 4):
        ws.cell(row=row_idx, column=1, value=t["id"]).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=2, value=t["open_date"]).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=3, value=t["ticker"]).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=4, value=t["sector"]).alignment = LEFT_ALIGN
        ws.cell(row=row_idx, column=5, value=t.get("signal_type", "default")).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=6, value=t["entry_actual"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=7, value=t["stop"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=8, value=t["target1"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=9, value=t["target2"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=10, value=t.get("score", 0)).number_format = "0.0"
        # Confidence not stored directly - use ml_probability if available
        conf = t.get("ml_probability")
        ws.cell(row=row_idx, column=11, value=f"{conf*100:.0f}%" if conf else "-").alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=12, value=t.get("days_open", 0)).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=13, value=t.get("current_price", t["entry_actual"])).number_format = "#,##0.00"
        
        pnl_cell = ws.cell(row=row_idx, column=14, value=t.get("unrealized_pnl_pct", 0))
        pnl_cell.number_format = "+0.00%;-0.00%"
        # Apply color based on pnl
        if t.get("unrealized_pnl_pct", 0) > 0:
            pnl_cell.fill = WIN_FILL
        elif t.get("unrealized_pnl_pct", 0) < 0:
            pnl_cell.fill = LOSS_FILL
        
        ws.cell(row=row_idx, column=15, value=t.get("mfe_pct", 0)).number_format = "+0.00%;-0.00%"
        ws.cell(row=row_idx, column=16, value=t.get("mae_pct", 0)).number_format = "+0.00%;-0.00%"
        
        # Status with emoji
        status = "🟡 Partial" if t.get("partial_closed") else "🟢 Active"
        status_cell = ws.cell(row=row_idx, column=17, value=status)
        status_cell.alignment = CENTER_ALIGN
    
    _autosize_columns(ws)
    ws.freeze_panes = "A4"


def build_closed_trades_sheet(ws, closed_trades):
    """Sheet 2: الصفقات المغلقة"""
    headers = [
        "Trade ID", "Open Date", "Close Date", "Ticker", "Sector", "Signal Type",
        "Entry", "Exit", "Days", "P&L %", "Result", "Exit Reason",
        "MFE%", "MAE%", "Score", "ADX"
    ]
    
    ws["A1"] = "📁 الصفقات المغلقة (Closed Trades)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"].alignment = CENTER_ALIGN
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        _style_header(cell)
    
    # Sort by close_date descending
    sorted_closed = sorted(closed_trades, key=lambda x: x.get("close_date", ""), reverse=True)
    
    for row_idx, t in enumerate(sorted_closed, 4):
        ws.cell(row=row_idx, column=1, value=t["id"]).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=2, value=t["open_date"]).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=3, value=t.get("close_date", "")).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=4, value=t["ticker"]).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=5, value=t["sector"]).alignment = LEFT_ALIGN
        ws.cell(row=row_idx, column=6, value=t.get("signal_type", "default")).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=7, value=t["entry_actual"]).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=8, value=t.get("exit_price", 0)).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=9, value=t.get("days_open", 0)).alignment = CENTER_ALIGN
        
        pnl_cell = ws.cell(row=row_idx, column=10, value=t.get("final_pnl_pct", 0) / 100)
        pnl_cell.number_format = "+0.00%;-0.00%"
        if "WIN" in t.get("result", ""):
            pnl_cell.fill = WIN_FILL
        elif "LOSS" in t.get("result", ""):
            pnl_cell.fill = LOSS_FILL
        else:
            pnl_cell.fill = NEUTRAL_FILL
        
        result_cell = ws.cell(row=row_idx, column=11, value=t.get("result", ""))
        result_cell.alignment = CENTER_ALIGN
        result_cell.font = Font(bold=True)
        if "WIN" in t.get("result", ""):
            result_cell.fill = WIN_FILL
        elif "LOSS" in t.get("result", ""):
            result_cell.fill = LOSS_FILL
        
        ws.cell(row=row_idx, column=12, value=t.get("exit_reason", "")).alignment = LEFT_ALIGN
        ws.cell(row=row_idx, column=13, value=t.get("mfe_pct", 0) / 100).number_format = "+0.00%;-0.00%"
        ws.cell(row=row_idx, column=14, value=t.get("mae_pct", 0) / 100).number_format = "+0.00%;-0.00%"
        ws.cell(row=row_idx, column=15, value=t.get("score", 0)).number_format = "0.0"
        ws.cell(row=row_idx, column=16, value=t.get("adx", 0)).number_format = "0.0"
    
    _autosize_columns(ws)
    ws.freeze_panes = "A4"


def build_stats_sheet(ws, stats):
    """Sheet 3: الإحصاءات الكلية"""
    ws["A1"] = "📊 إحصاءات الأداء الإجمالي"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = CENTER_ALIGN
    
    ws["A2"] = f"محسوب في: {stats.get('computed_at', '')[:19]}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:D2")
    ws["A2"].alignment = CENTER_ALIGN
    
    # KPIs
    row = 4
    kpis = [
        ("إجمالي الصفقات", stats.get("total_trades", 0), ""),
        ("الصفقات النشطة", stats.get("active_trades", 0), ""),
        ("الفائزة", stats.get("wins", 0), ""),
        ("الخاسرة", stats.get("losses", 0), ""),
        ("Win Rate", f"{stats.get('win_rate_pct', 0)}%", "win"),
        ("متوسط الربح", f"+{stats.get('avg_win_pct', 0)}%", "win"),
        ("متوسط الخسارة", f"{stats.get('avg_loss_pct', 0)}%", "loss"),
        ("Profit Factor", stats.get("profit_factor", "N/A"), "neutral"),
        ("إجمالي P&L%", f"{stats.get('total_pnl_pct', 0)}%", "win" if stats.get('total_pnl_pct', 0) > 0 else "loss"),
    ]
    
    ws.cell(row=row, column=1, value="المقياس").fill = HEADER_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).alignment = CENTER_ALIGN
    ws.cell(row=row, column=2, value="القيمة").fill = HEADER_FILL
    ws.cell(row=row, column=2).font = HEADER_FONT
    ws.cell(row=row, column=2).alignment = CENTER_ALIGN
    
    for label, value, fill_type in kpis:
        row += 1
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=value)
        c.alignment = CENTER_ALIGN
        if fill_type == "win":
            c.fill = WIN_FILL
        elif fill_type == "loss":
            c.fill = LOSS_FILL
        elif fill_type == "neutral":
            c.fill = NEUTRAL_FILL
    
    # Best / Worst
    row += 2
    ws.cell(row=row, column=1, value="🏆 أفضل صفقة").font = Font(bold=True, size=12)
    row += 1
    best = stats.get("best_trade") or {}
    ws.cell(row=row, column=1, value=f"{best.get('ticker','-')} ({best.get('date','-')})")
    ws.cell(row=row, column=2, value=f"+{best.get('pnl', 0):.2f}%").fill = WIN_FILL
    
    row += 2
    ws.cell(row=row, column=1, value="📉 أسوأ صفقة").font = Font(bold=True, size=12)
    row += 1
    worst = stats.get("worst_trade") or {}
    ws.cell(row=row, column=1, value=f"{worst.get('ticker','-')} ({worst.get('date','-')})")
    ws.cell(row=row, column=2, value=f"{worst.get('pnl', 0):.2f}%").fill = LOSS_FILL
    
    # By Signal Type
    row += 3
    ws.cell(row=row, column=1, value="📈 الأداء حسب نوع الإشارة").font = Font(bold=True, size=14, color="1F4E78")
    row += 2
    headers = ["Signal Type", "Total", "Wins", "Win Rate", "Avg P&L"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        _style_subheader(cell)
    
    for st, d in (stats.get("by_signal_type") or {}).items():
        row += 1
        ws.cell(row=row, column=1, value=st)
        ws.cell(row=row, column=2, value=d["total"]).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3, value=d["wins"]).alignment = CENTER_ALIGN
        wr_cell = ws.cell(row=row, column=4, value=f"{d['win_rate']}%")
        wr_cell.alignment = CENTER_ALIGN
        if d["win_rate"] >= 60:
            wr_cell.fill = WIN_FILL
        elif d["win_rate"] < 40:
            wr_cell.fill = LOSS_FILL
        pnl_cell = ws.cell(row=row, column=5, value=f"{d['avg_pnl']:+.2f}%")
        pnl_cell.alignment = CENTER_ALIGN
        if d["avg_pnl"] > 0:
            pnl_cell.fill = WIN_FILL
        else:
            pnl_cell.fill = LOSS_FILL
    
    # By Sector
    row += 3
    ws.cell(row=row, column=1, value="🏢 الأداء حسب القطاع").font = Font(bold=True, size=14, color="1F4E78")
    row += 2
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        _style_subheader(cell)
    
    for sec, d in (stats.get("by_sector") or {}).items():
        row += 1
        ws.cell(row=row, column=1, value=sec)
        ws.cell(row=row, column=2, value=d["total"]).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3, value=d["wins"]).alignment = CENTER_ALIGN
        wr_cell = ws.cell(row=row, column=4, value=f"{d['win_rate']}%")
        wr_cell.alignment = CENTER_ALIGN
        if d["win_rate"] >= 60:
            wr_cell.fill = WIN_FILL
        elif d["win_rate"] < 40:
            wr_cell.fill = LOSS_FILL
        pnl_cell = ws.cell(row=row, column=5, value=f"{d['avg_pnl']:+.2f}%")
        pnl_cell.alignment = CENTER_ALIGN
    
    # Tired stocks
    row += 3
    ws.cell(row=row, column=1, value="⚠️ الأسهم المُتعِبة (Win Rate < 30% مع 3+ صفقات)").font = Font(bold=True, size=12, color="C00000")
    row += 2
    if stats.get("tired_stocks"):
        for tk in stats["tired_stocks"]:
            row += 1
            c = ws.cell(row=row, column=1, value=f"{tk['ticker']}: {tk['wins']}/{tk['total']} = {tk['win_rate']}%")
            c.fill = LOSS_FILL
    else:
        ws.cell(row=row, column=1, value="(لا توجد بعد - يحتاج 3+ صفقات لكل سهم)").font = Font(italic=True, color="666666")
    
    _autosize_columns(ws, max_width=40)


def build_per_stock_sheet(ws, closed_trades):
    """Sheet 4: أداء كل سهم"""
    ws["A1"] = "📊 أداء كل سهم (Per-Stock Performance)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = CENTER_ALIGN
    
    headers = ["Ticker", "Sector", "Total", "Wins", "Losses", "Win Rate", "Avg P&L%", "Best", "Worst"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        _style_header(cell)
    
    by_stock = defaultdict(lambda: {"sector": "", "trades": []})
    for t in closed_trades:
        by_stock[t["ticker"]]["sector"] = t.get("sector", "?")
        by_stock[t["ticker"]]["trades"].append(t)
    
    rows = []
    for tk, d in by_stock.items():
        trades = d["trades"]
        total = len(trades)
        wins = sum(1 for x in trades if "WIN" in x.get("result", ""))
        losses = sum(1 for x in trades if "LOSS" in x.get("result", ""))
        avg_pnl = sum(x["final_pnl_pct"] for x in trades) / total
        best = max(trades, key=lambda x: x["final_pnl_pct"])["final_pnl_pct"]
        worst = min(trades, key=lambda x: x["final_pnl_pct"])["final_pnl_pct"]
        
        rows.append({
            "ticker": tk, "sector": d["sector"], "total": total,
            "wins": wins, "losses": losses,
            "win_rate": wins / total * 100 if total else 0,
            "avg_pnl": avg_pnl, "best": best, "worst": worst,
        })
    
    rows.sort(key=lambda x: -x["win_rate"])
    
    for row_idx, r in enumerate(rows, 4):
        ws.cell(row=row_idx, column=1, value=r["ticker"]).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=2, value=r["sector"]).alignment = LEFT_ALIGN
        ws.cell(row=row_idx, column=3, value=r["total"]).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=4, value=r["wins"]).alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=5, value=r["losses"]).alignment = CENTER_ALIGN
        wr_cell = ws.cell(row=row_idx, column=6, value=f"{r['win_rate']:.1f}%")
        wr_cell.alignment = CENTER_ALIGN
        if r["win_rate"] >= 60:
            wr_cell.fill = WIN_FILL
        elif r["win_rate"] < 40 and r["total"] >= 3:
            wr_cell.fill = LOSS_FILL
        ws.cell(row=row_idx, column=7, value=f"{r['avg_pnl']:+.2f}%").alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=8, value=f"{r['best']:+.2f}%").alignment = CENTER_ALIGN
        ws.cell(row=row_idx, column=9, value=f"{r['worst']:+.2f}%").alignment = CENTER_ALIGN
    
    _autosize_columns(ws)
    ws.freeze_panes = "A4"


def build_dashboard(output_dir=None, today_str=None):
    """البناء الرئيسي للـ Excel dashboard."""
    if output_dir is None:
        output_dir = PAPER_DIR
    if today_str is None:
        today_str = datetime.now().strftime("%Y-%m-%d")
    
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Load data
    if not F_TRADES.exists():
        print("  ⚠️ لا توجد paper_trades.json بعد - تخطي بناء Excel")
        return None
    
    with open(F_TRADES, "r", encoding="utf-8") as f:
        db = json.load(f)
    
    # Stats
    from paper_trading_engine import compute_stats
    stats = compute_stats()
    
    # Build workbook
    wb = Workbook()
    
    # Sheet 1: Active
    ws1 = wb.active
    ws1.title = "Active Trades"
    build_active_trades_sheet(ws1, db.get("active", []))
    
    # Sheet 2: Closed
    ws2 = wb.create_sheet("Closed Trades")
    build_closed_trades_sheet(ws2, db.get("closed", []))
    
    # Sheet 3: Stats
    ws3 = wb.create_sheet("Performance Stats")
    build_stats_sheet(ws3, stats)
    
    # Sheet 4: Per-Stock
    ws4 = wb.create_sheet("Per-Stock")
    build_per_stock_sheet(ws4, db.get("closed", []))
    
    # Save
    output_path = output_dir / f"dashboard_{today_str}.xlsx"
    wb.save(output_path)
    
    # Also save as latest.xlsx for easy access
    latest_path = output_dir / "latest.xlsx"
    wb.save(latest_path)
    
    print(f"  ✓ Dashboard: {output_path}")
    print(f"  ✓ Latest:    {latest_path}")
    
    return output_path


if __name__ == "__main__":
    build_dashboard()
